from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


EXCEL_THEMES = {
    "light": {
        "background": "FFFFFFFF",
        "alternate_background": "FFF7F7F7",
        "text": "FF000000",
        "grid": "FFD0D0D0",
        "header_background": "FF007FFF",
        "header_text": "FFFFFFFF",
        "calendar_occupied_fill": "FFC6EFCE",
    },
    "dark": {
        "background": "FF1E1E1E",
        "alternate_background": "FF2A2A2A",
        "text": "FFF2F2F2",
        "grid": "FF555555",
        "header_background": "FF6395EE",
        "header_text": "FFFFFFFF",
        "calendar_occupied_fill": "FF2F6B3B",
    },
}

VACATION_COLOR = "FF800080"
TRAINING_COLOR = "FFFFA500"
REQUESTED_DAYS_OFF_COLOR = "FFFF1493"
WHITE = "FFFFFFFF"
BLACK = "FF000000"


def normalize_theme_name(value: str | None) -> str:
    name = str(value or "light").strip().lower()
    return name if name in EXCEL_THEMES else "light"


def normalize_column_key(value) -> str:
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def normalize_date(value) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    try:
        converted = pd.to_datetime(value, errors="coerce")
    except Exception:
        return None

    if pd.isna(converted):
        return None

    return converted.date()


def normalize_date_ranges(ranges) -> list[tuple[date, date]]:
    if not ranges:
        return []

    normalized: list[tuple[date, date]] = []

    for item in ranges:
        if isinstance(item, dict):
            start = item.get("start")
            end = item.get("end")
        else:
            try:
                start, end = item
            except Exception:
                continue

        start_date = normalize_date(start)
        end_date = normalize_date(end)

        if start_date is None or end_date is None:
            continue

        if end_date < start_date:
            start_date, end_date = end_date, start_date

        normalized.append((start_date, end_date))

    return normalized


def normalize_date_list(values) -> set[date]:
    if values is None:
        return set()

    if isinstance(values, (str, date, datetime, pd.Timestamp)):
        values = [values]

    result: set[date] = set()
    for value in values:
        normalized = normalize_date(value)
        if normalized is not None:
            result.add(normalized)
    return result


def date_ranges_to_date_set(ranges) -> set[date]:
    result: set[date] = set()

    for start, end in ranges:
        current = start
        while current <= end:
            result.add(current)
            current += timedelta(days=1)

    return result


def format_calendar_header(value: date) -> str:
    return f"{value:%a}, {value:%b} {value.day}"


def date_in_any_range(value: date, ranges) -> bool:
    return any(start <= value <= end for start, end in ranges)


def is_blank(value) -> bool:
    if value is None or value == "":
        return True

    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def is_line_number_column(value) -> bool:
    return normalize_column_key(value) in {"linenumber", "linenumbers"}


def is_extra_vacation_column(value) -> bool:
    return "vacation" in normalize_column_key(value)


def is_training_column(value) -> bool:
    return "training" in normalize_column_key(value)


def is_requested_days_off_column(value) -> bool:
    key = normalize_column_key(value)
    return (
        "requested" in key
        and (
            "dayoff" in key
            or "daysoff" in key
            or "dateoff" in key
            or "datesoff" in key
            or "days" in key
            or "dates" in key
        )
    )


def _make_unique_excel_headers(headers) -> list[str]:
    """Excel tables require non-empty, unique string headers."""
    used: set[str] = set()
    result: list[str] = []

    for idx, header in enumerate(headers, start=1):
        base = "" if header is None else str(header).strip()
        if not base:
            base = f"Column{idx}"

        base = base[:240]
        candidate = base
        counter = 2

        while candidate in used:
            suffix = f"_{counter}"
            candidate = f"{base[:240 - len(suffix)]}{suffix}"
            counter += 1

        used.add(candidate)
        result.append(candidate)

    return result


def _sanitize_table_name(name) -> str:
    cleaned = re.sub(r"\W+", "_", str(name))
    if not cleaned:
        cleaned = "Table1"
    if cleaned[0].isdigit():
        cleaned = "_" + cleaned
    return cleaned[:255]


def _visible_text_length(value) -> int:
    if value is None:
        return 0
    return max((len(part) for part in str(value).splitlines()), default=0)


def _autosize_column(ws, col_idx: int, *, min_width: float, max_width: float) -> None:
    max_len = 0
    for row in range(1, ws.max_row + 1):
        max_len = max(max_len, _visible_text_length(ws.cell(row=row, column=col_idx).value))

    width = min(max(max_len + 2, min_width), max_width)
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_border_side(cell, *, left=None, right=None) -> None:
    old = cell.border
    cell.border = Border(
        left=left if left is not None else copy(old.left),
        right=right if right is not None else copy(old.right),
        top=copy(old.top),
        bottom=copy(old.bottom),
        diagonal=copy(old.diagonal),
        diagonal_direction=old.diagonal_direction,
        diagonalUp=old.diagonalUp,
        diagonalDown=old.diagonalDown,
        outline=old.outline,
        vertical=copy(old.vertical),
        horizontal=copy(old.horizontal),
    )


def _column_width_for_position(column_widths, position: int, original_column) -> Optional[float]:
    if column_widths is None:
        return None

    value = None

    if isinstance(column_widths, Mapping):
        if position in column_widths:
            value = column_widths[position]
        elif original_column in column_widths:
            value = column_widths[original_column]
        elif str(original_column) in column_widths:
            value = column_widths[str(original_column)]
    elif isinstance(column_widths, Sequence) and not isinstance(column_widths, (str, bytes)):
        if position < len(column_widths):
            value = column_widths[position]

    if value is None:
        return None

    try:
        return min(255.0, max(2.0, float(value)))
    except (TypeError, ValueError):
        return None


def export_master_lines_to_excel_table(
    df: pd.DataFrame,
    output_path: str | Path,
    *,
    sheet_name: str = "Master Lines",
    table_name: str = "MasterLinesTable",
    calendar_cols=None,
    training_start=None,
    training_end=None,
    vacation_ranges=None,
    requested_days_off_dates=None,
    requested_days_off_ranges=None,
    theme: str = "light",
    table_style: str | None = "TableStyleLight9",
    calendar_col_width: float = 32,
    calendar_row_height: float = 40,
    header_row_height: float = 45,
    non_calendar_max_width: float = 32,
    non_calendar_min_width: float = 8,
    body_font_size: float = 12,
    header_font_size: float = 14,
    column_widths=None,
    sheet_zoom: int = 100,
    freeze_panes: str | None = "A2",
) -> Path:
    """
    Export a DataFrame with formatting that mirrors BidSpreadsheetViewer.

    The DataFrame passed here is exported exactly as supplied. Therefore:
      - direct export after sorting: pass the sorted DataFrame;
      - export from the visualizer: pass viewer.get_export_dataframe(), which
        already contains the current row order and excludes hidden rows/columns.

    New visualizer-matching features:
      - light/dark theme colors;
      - blue normal headers;
      - orange training headers/markers;
      - purple vacation headers/markers;
      - pink requested-days-off headers/markers;
      - green occupied calendar cells;
      - alternating body rows and grid borders;
      - bold Line Number cells;
      - optional current-view column widths and Excel zoom.

    ``column_widths`` may be a sequence aligned with ``df.columns`` or a mapping
    keyed by zero-based position, original column object, or string column name.
    """
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df must be a pandas DataFrame")

    if df.shape[1] == 0:
        raise ValueError("Cannot export a DataFrame with no visible columns.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    theme_name = normalize_theme_name(theme)
    colors = EXCEL_THEMES[theme_name]

    training_start = normalize_date(training_start)
    training_end = normalize_date(training_end)
    vacation_ranges = normalize_date_ranges(vacation_ranges)
    requested_ranges = normalize_date_ranges(requested_days_off_ranges)
    requested_dates = normalize_date_list(requested_days_off_dates)
    requested_dates.update(date_ranges_to_date_set(requested_ranges))

    export_df = df.copy()
    original_columns = list(export_df.columns)

    if calendar_cols is None:
        calendar_date_set = {
            normalized
            for col in original_columns
            if (normalized := normalize_date(col)) is not None
        }
    else:
        calendar_date_set = {
            normalized
            for col in calendar_cols
            if (normalized := normalize_date(col)) is not None
        }

    preliminary_headers: list[object] = []
    date_by_position: dict[int, date] = {}

    for position, column in enumerate(original_columns):
        column_date = normalize_date(column)
        if column_date in calendar_date_set:
            preliminary_headers.append(format_calendar_header(column_date))
            date_by_position[position] = column_date
        else:
            preliminary_headers.append(column)

    final_headers = _make_unique_excel_headers(preliminary_headers)
    export_df.columns = final_headers

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(output_path)
    worksheet = workbook[sheet_name]

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    worksheet.freeze_panes = freeze_panes
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = max(10, min(400, int(sheet_zoom)))
    worksheet.sheet_view.zoomScaleNormal = 100

    # Keep the real Excel table/filter behavior from the original exporter.
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=_sanitize_table_name(table_name), ref=table_ref)

    if table_style:
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,  # body striping is painted manually below
            showColumnStripes=False,
        )

    worksheet.add_table(table)

    normal_fill = PatternFill("solid", fgColor=colors["background"])
    alternate_fill = PatternFill("solid", fgColor=colors["alternate_background"])
    calendar_fill = PatternFill("solid", fgColor=colors["calendar_occupied_fill"])
    normal_header_fill = PatternFill("solid", fgColor=colors["header_background"])
    vacation_fill = PatternFill("solid", fgColor=VACATION_COLOR)
    training_fill = PatternFill("solid", fgColor=TRAINING_COLOR)
    requested_fill = PatternFill("solid", fgColor=REQUESTED_DAYS_OFF_COLOR)

    grid_side = Side(style="thin", color=colors["grid"])
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)

    training_start_side = Side(style="thick", color=TRAINING_COLOR)
    training_end_side = Side(style="mediumDashed", color=TRAINING_COLOR)
    requested_start_side = Side(style="thick", color=REQUESTED_DAYS_OFF_COLOR)
    requested_end_side = Side(style="mediumDashed", color=REQUESTED_DAYS_OFF_COLOR)
    vacation_start_side = Side(style="thick", color=VACATION_COLOR)
    vacation_end_side = Side(style="mediumDashed", color=VACATION_COLOR)

    date_to_excel_col = {column_date: position + 1 for position, column_date in date_by_position.items()}
    calendar_excel_cols = set(date_to_excel_col.values())

    first_bid_date = min(date_to_excel_col, default=None)
    last_bid_date = max(date_to_excel_col, default=None)

    worksheet.row_dimensions[1].height = header_row_height

    # Base body formatting first. Specific calendar fills are applied afterward.
    for row in range(2, max_row + 1):
        worksheet.row_dimensions[row].height = calendar_row_height
        row_fill = normal_fill if row % 2 == 0 else alternate_fill

        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.fill = row_fill
            cell.border = grid_border
            cell.font = Font(color=colors["text"], size=body_font_size)
            cell.alignment = Alignment(vertical="center", horizontal="left")

    # Header formatting and special metric/date colors.
    for position, final_header in enumerate(final_headers):
        excel_col = position + 1
        original_column = original_columns[position]
        column_date = date_by_position.get(position)
        header_cell = worksheet.cell(row=1, column=excel_col)

        header_fill = normal_header_fill
        header_font_color = colors["header_text"]

        # Same precedence as the visualizer: training > requested > vacation.
        uses_training = is_training_column(original_column)
        if column_date is not None and training_start is not None and training_end is not None:
            training_range_start = min(training_start, training_end)
            training_range_end = max(training_start, training_end)
            uses_training = uses_training or training_range_start <= column_date <= training_range_end

        uses_requested = is_requested_days_off_column(original_column)
        if column_date is not None:
            uses_requested = uses_requested or column_date in requested_dates

        uses_vacation = is_extra_vacation_column(original_column)
        if column_date is not None:
            uses_vacation = uses_vacation or date_in_any_range(column_date, vacation_ranges)

        if uses_training:
            header_fill = training_fill
            header_font_color = BLACK
        elif uses_requested:
            header_fill = requested_fill
            header_font_color = WHITE
        elif uses_vacation:
            header_fill = vacation_fill
            header_font_color = WHITE

        header_cell.fill = header_fill
        header_cell.font = Font(color=header_font_color, bold=False, size=header_font_size)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = grid_border

        if is_line_number_column(original_column):
            for row in range(2, max_row + 1):
                worksheet.cell(row=row, column=excel_col).font = Font(
                    color=colors["text"],
                    size=body_font_size,
                    bold=True,
                )

        if column_date is not None:
            for row in range(2, max_row + 1):
                cell = worksheet.cell(row=row, column=excel_col)
                if not is_blank(cell.value):
                    cell.fill = calendar_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Match the left/right marker precedence used by DataFrameTableModel.
    for column_date, excel_col in date_to_excel_col.items():
        left_side = None
        right_side = None

        if training_start is not None and column_date == training_start:
            left_side = training_start_side
        if training_end is not None and column_date == training_end:
            right_side = training_end_side

        for requested_start, requested_end in requested_ranges:
            if column_date == requested_start and left_side is None:
                left_side = requested_start_side
            if column_date == requested_end and right_side is None:
                right_side = requested_end_side

        if column_date in requested_dates and not date_in_any_range(column_date, requested_ranges):
            if left_side is None:
                left_side = requested_start_side
            if right_side is None:
                right_side = requested_end_side

        for vacation_start, vacation_end in vacation_ranges:
            if column_date == vacation_start and left_side is None:
                left_side = vacation_start_side
            if column_date == vacation_end and right_side is None:
                right_side = vacation_end_side

        if left_side is not None or right_side is not None:
            for row in range(1, max_row + 1):
                _add_border_side(
                    worksheet.cell(row=row, column=excel_col),
                    left=left_side,
                    right=right_side,
                )

    # Preserve the original vacation markers immediately outside the bid period.
    for vacation_start, vacation_end in vacation_ranges:
        if first_bid_date is not None and vacation_end == first_bid_date - timedelta(days=1):
            excel_col = date_to_excel_col[first_bid_date]
            for row in range(1, max_row + 1):
                _add_border_side(worksheet.cell(row=row, column=excel_col), left=vacation_end_side)

        if last_bid_date is not None and vacation_start == last_bid_date + timedelta(days=1):
            excel_col = date_to_excel_col[last_bid_date]
            for row in range(1, max_row + 1):
                _add_border_side(worksheet.cell(row=row, column=excel_col), right=vacation_start_side)

    # Use visualizer-provided widths when available; otherwise keep the old defaults.
    for position, original_column in enumerate(original_columns):
        excel_col = position + 1
        explicit_width = _column_width_for_position(column_widths, position, original_column)

        if explicit_width is not None:
            worksheet.column_dimensions[get_column_letter(excel_col)].width = explicit_width
        elif excel_col in calendar_excel_cols:
            worksheet.column_dimensions[get_column_letter(excel_col)].width = calendar_col_width
        else:
            _autosize_column(
                worksheet,
                excel_col,
                min_width=non_calendar_min_width,
                max_width=non_calendar_max_width,
            )

    workbook.save(output_path)
    return output_path
