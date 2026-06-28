import re
import pandas as pd
from copy import copy
from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def export_master_lines_to_excel_table(
    df,
    output_path,
    *,
    sheet_name="Master Lines",
    table_name="MasterLinesTable",
    calendar_cols=None,
    training_start=None,
    training_end=None,
    vacation_ranges=None,
    table_style="TableStyleLight9",
    calendar_col_width=32,
    calendar_row_height=28,
    header_row_height=36,
    non_calendar_max_width=32,
):
    """
    Exports a pandas DataFrame to Excel as a real Excel table.

    Calendar formatting:
        - Calendar headers are displayed like: Wed, May 27
        - Any non-empty calendar cell is highlighted green.
        - Training start date gets thick solid red LEFT border.
        - Training end date gets dashed red RIGHT border.
        - Vacation start date gets thick solid royal blue LEFT border.
        - Vacation end date gets dashed royal blue RIGHT border.
        - Vacation date headers are royal blue.
        - Training date headers are red.
        - Training header color wins if it overlaps vacation.
        - Calendar columns all get the same fixed width.
        - Data rows all get the same fixed height.

    vacation_ranges format:
        vacation_ranges=[
            {"start": "2023-05-01", "end": "2023-05-21"},
            {"start": "2023-06-01", "end": "2023-06-08"},
        ]
    """

    df = df.copy()

    def normalize_date(value):
        if value is None or value == "":
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        try:
            return pd.to_datetime(value).date()
        except Exception:
            return None

    def format_calendar_header(d):
        # Cross-platform. Avoids "%-d", which does not work reliably on Windows.
        return f"{d:%a}, {d:%b} {d.day}"

    def sanitize_table_name(name):
        name = re.sub(r"\W+", "_", str(name))
        if not name:
            name = "Table1"
        if name[0].isdigit():
            name = "_" + name
        return name

    def normalize_date_ranges(ranges):
        """
        Accepts:
            [{"start": "2023-05-01", "end": "2023-05-21"}]

        Also accepts:
            [("2023-05-01", "2023-05-21")]
        """

        if not ranges:
            return []

        normalized = []

        for item in ranges:
            if isinstance(item, dict):
                start = item.get("start")
                end = item.get("end")
            else:
                try:
                    start, end = item
                except Exception:
                    continue

            start_date = normalize_date(start)
            end_date = normalize_date(end)

            if start_date is None or end_date is None:
                continue

            if end_date < start_date:
                start_date, end_date = end_date, start_date

            normalized.append((start_date, end_date))

        return normalized

    def visible_text_length(value):
        if value is None:
            return 0

        text = str(value)

        if "\n" in text:
            return max(len(part) for part in text.splitlines())

        return len(text)

    def autosize_column(ws, col_idx, min_width=8, max_width=30):
        max_len = 0

        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            max_len = max(max_len, visible_text_length(value))

        width = max(min_width, max_len + 2)
        width = min(width, max_width)

        ws.column_dimensions[get_column_letter(col_idx)].width = width

    def add_border_side(cell, *, left=None, right=None):
        old = cell.border

        cell.border = Border(
            left=left if left is not None else copy(old.left),
            right=right if right is not None else copy(old.right),
            top=copy(old.top),
            bottom=copy(old.bottom),
            diagonal=copy(old.diagonal),
            diagonal_direction=old.diagonal_direction,
            diagonalUp=old.diagonalUp,
            diagonalDown=old.diagonalDown,
            outline=old.outline,
            vertical=copy(old.vertical),
            horizontal=copy(old.horizontal),
        )

    def date_in_any_range(d, ranges):
        return any(start <= d <= end for start, end in ranges)

    training_start = normalize_date(training_start)
    training_end = normalize_date(training_end)
    vacation_ranges = normalize_date_ranges(vacation_ranges)

    # Determine calendar columns.
    if calendar_cols is None:
        calendar_date_set = {
            normalize_date(col)
            for col in df.columns
            if normalize_date(col) is not None
        }
    else:
        calendar_date_set = {
            normalize_date(col)
            for col in calendar_cols
            if normalize_date(col) is not None
        }

    # Rename calendar columns to visible headers like "Wed, May 27".
    new_columns = []
    display_header_to_date = {}

    for col in df.columns:
        col_date = normalize_date(col)

        if col_date in calendar_date_set:
            display_header = format_calendar_header(col_date)

            new_columns.append(display_header)
            display_header_to_date[display_header] = col_date
        else:
            new_columns.append(str(col))

    df.columns = new_columns

    # Write DataFrame
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = "A2"

    # Create Excel table
    table_name = sanitize_table_name(table_name)
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

    tab = Table(displayName=table_name, ref=table_ref)

    style = TableStyleInfo(
        name=table_style,  # TableStyleLight9
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Helps some spreadsheet programs recognize the filterable range.
    ws.auto_filter.ref = table_ref

    # Fills
    green_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC6EFCE",
    )

    vacation_header_fill = PatternFill(
        fill_type="solid",
        fgColor="FF800080",  # Royal blue
    )

    training_header_fill = PatternFill(
        fill_type="solid",
        fgColor="FFFFA500",  # Red
    )

    white_bold_font = Font(
        color="FFFFFFFF",
        bold=True,
    )

    # Borders
    training_start_side = Side(
        style="thick",
        color="FFFFA500",
    )

    training_end_side = Side(
        style="mediumDashed",
        color="FFFFA500",
    )

    vacation_start_side = Side(
        style="thick",
        color="FF800080",
    )

    vacation_end_side = Side(
        style="mediumDashed",
        color="FF800080",
    )

    # Build real-date -> Excel column index map.
    date_to_excel_col = {}

    for col_idx in range(1, max_col + 1):
        header_value = ws.cell(row=1, column=col_idx).value

        if header_value in display_header_to_date:
            real_date = display_header_to_date[header_value]
            date_to_excel_col[real_date] = col_idx

    calendar_excel_cols = set(date_to_excel_col.values())

    if date_to_excel_col:
        first_bid_date = min(date_to_excel_col)
        last_bid_date = max(date_to_excel_col)
    else:
        first_bid_date = None
        last_bid_date = None

    # Header row height
    ws.row_dimensions[1].height = header_row_height

    # Basic header formatting
    for cell in ws[1]:
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # Fixed row height for data rows.
    # Excel row height applies to the whole row.
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = calendar_row_height

    # Calendar cell formatting
    for excel_col in calendar_excel_cols:
        col_letter = get_column_letter(excel_col)
        ws.column_dimensions[col_letter].width = calendar_col_width

        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=excel_col)

            if cell.value not in (None, ""):
                cell.fill = green_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    # Color vacation headers royal blue
    for d, excel_col in date_to_excel_col.items():
        if date_in_any_range(d, vacation_ranges):
            header_cell = ws.cell(row=1, column=excel_col)
            header_cell.fill = vacation_header_fill
            header_cell.font = white_bold_font

    # Color training headers red.
    # This is done after vacation so training wins if there is overlap.
    if training_start is not None and training_end is not None:
        start = min(training_start, training_end)
        end = max(training_start, training_end)

        for d, excel_col in date_to_excel_col.items():
            if start <= d <= end:
                header_cell = ws.cell(row=1, column=excel_col)
                header_cell.fill = training_header_fill
                header_cell.font = white_bold_font

    # Training start: thick solid red LEFT border
    if training_start in date_to_excel_col:
        excel_col = date_to_excel_col[training_start]

        for row in range(1, max_row + 1):
            add_border_side(
                ws.cell(row=row, column=excel_col),
                left=training_start_side,
            )

    # Training end: dashed red RIGHT border
    if training_end in date_to_excel_col:
        excel_col = date_to_excel_col[training_end]

        for row in range(1, max_row + 1):
            add_border_side(
                ws.cell(row=row, column=excel_col),
                right=training_end_side,
            )

    # Vacation start/end borders inside the bid period
    for vacation_start, vacation_end in vacation_ranges:

        # Vacation start inside bid period:
        # thick solid royal blue LEFT border
        if vacation_start in date_to_excel_col:
            excel_col = date_to_excel_col[vacation_start]

            for row in range(1, max_row + 1):
                add_border_side(
                    ws.cell(row=row, column=excel_col),
                    left=vacation_start_side,
                )

        # Vacation end inside bid period:
        # dashed royal blue RIGHT border
        if vacation_end in date_to_excel_col:
            excel_col = date_to_excel_col[vacation_end]

            for row in range(1, max_row + 1):
                add_border_side(
                    ws.cell(row=row, column=excel_col),
                    right=vacation_end_side,
                )

        # Special case:
        # Vacation ended the day before the bid period starts.
        # Put dashed royal blue LEFT border on first bid date.
        if first_bid_date is not None:
            if vacation_end == first_bid_date - timedelta(days=1):
                excel_col = date_to_excel_col[first_bid_date]

                for row in range(1, max_row + 1):
                    add_border_side(
                        ws.cell(row=row, column=excel_col),
                        left=vacation_end_side,
                    )

        # Special case:
        # Vacation starts the day after the bid period ends.
        # Put thick solid royal blue RIGHT border on last bid date.
        if last_bid_date is not None:
            if vacation_start == last_bid_date + timedelta(days=1):
                excel_col = date_to_excel_col[last_bid_date]

                for row in range(1, max_row + 1):
                    add_border_side(
                        ws.cell(row=row, column=excel_col),
                        right=vacation_start_side,
                    )

    # Auto-size only the non-calendar columns
    for col_idx in range(1, max_col + 1):
        if col_idx in calendar_excel_cols:
            continue

        autosize_column(
            ws,
            col_idx,
            min_width=8,
            max_width=non_calendar_max_width,
        )

    wb.save(output_path)